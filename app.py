import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

# Set page configuration
st.set_page_config(
    page_title="Unbilled WIP Report Generator",
    page_icon="📊",
    layout="wide"
)

# Load and display logo
logo = Image.open("image.png")
st.image(logo, width=100)

# Markdown header
st.markdown("""
# 📊 Unbilled WIP Report Generator
Welcome! Upload your Fast Track Excel file to generate a clean report.
""")

# Required headers
required_headers = [
    "Brand", "Timesheet ID", "Timesheet Code", "Client Ref", "Client Name",
    "Invoice Group", "Interpreter Status", "Purchase Order", "Job Order ID",
    "Week ending date", "Contractor Name", "Bill Rate Description", "Bill Units",
    "Bill Rate", "Total Bill", "Work Location", "Business Unit", "Job Description",
    "Project Code1"
]

# Invoice Group filters
experis_groups = {
    "T3-4-PO", "EB-4-NO PO", "EB-4-PO", "EB-CalendarMonthly-PO",
    "EB-M-No PO", "EB-M-PO", "EB-W-No PO", "EB-W-PO",
    "T3-4-ONLI", "T3-4-SCHE", "T3-M-No PO", "T3-M-PO",
    "T3-SelfBIll-NONPO", "T3-W-Stand", "TCS self bill", "T3-W-PO-Ariba"
}

manpower_groups = {
    "TCS Weekly-Consolidated-PO", "TCS Consolidated-W- PO", "TCS weekly PO", "TCS EB-W- PO",
    "TCS -Weekly- Consolidated- No PO - 560 Back up"
}

# File uploader
uploaded_file = st.file_uploader("Upload Fast Track Excel File (.xlsx)", type=["xlsx"])

if uploaded_file:
    df_raw = pd.read_excel(uploaded_file, engine="openpyxl")

    for header in required_headers:
        if header not in df_raw.columns:
            df_raw[header] = ""

    def determine_brand(inv_group):
        if inv_group in experis_groups:
            return "Experis"
        elif inv_group in manpower_groups:
            return "Talent Solutions" if "560" in inv_group else "Manpower"
        return ""

    df_raw["Brand"] = df_raw["Invoice Group"].apply(determine_brand)
    df_raw["Week ending date"] = pd.to_datetime(df_raw["Week ending date"], errors='coerce')
    df_raw = df_raw[df_raw["Week ending date"].notna()]
    df_raw.sort_values(by=["Client Name", "Contractor Name", "Week ending date"], inplace=True)
    df_raw["Week ending date"] = df_raw["Week ending date"].dt.strftime('%m/%d/%Y')
    df_raw = df_raw[required_headers]

    df_experis = df_raw[df_raw["Brand"] == "Experis"]
    df_manpower = df_raw[df_raw["Brand"].isin(["Manpower", "Talent Solutions"])]

    output = io.BytesIO()
    current_time = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    file_name = f"Unbilled WIP Report - {current_time}.xlsx"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_raw.to_excel(writer, sheet_name="All", index=False)
        df_experis.to_excel(writer, sheet_name="Experis", index=False)
        df_manpower.to_excel(writer, sheet_name="Manpower", index=False)

    output.seek(0)
    wb = load_workbook(output)

    for sheet_name in ["Experis", "Manpower"]:
        ws = wb[sheet_name]
        header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        week_col_index = None

        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value == "Week ending date":
                week_col_index = cell.column

        if week_col_index:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.column == week_col_index and cell.value:
                        cell.number_format = "MM/DD/YYYY"

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.success("✅ Report generated successfully!")
    st.download_button("📥 Download Unbilled WIP Report", final_output.getvalue(), file_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")